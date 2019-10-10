from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image


class ImageToExcel(object):

    def __init__(self, image_name: str = "img.png", excel_name: str = "picture.xlsx"):
        self.image_name = image_name
        self.excel_name = excel_name

    def __get_pixels(self, image_name: str = None):
        """
        Переводит изоображение в пиксели

        :param image_name:
        :return:
        """

        if image_name is None:
            image_name = self.image_name
        img = Image.open(image_name, 'r')
        w, h = img.size
        pixel = list(img.getdata())
        return [pixel[item:item + w] for item in range(0, w * h, w)]

    def __clamp(self, x: int):
        """
        Возвращет число

        :param x:
        :return:
        """

        return max(0, min(x, 255))

    def __get_hex_color(self, r: int, g: int, b: int):
        """
        Преводит RGB в HEX

        :param r:
        :param g:
        :param b:
        :return:
        """

        return f"{self.__clamp(r):02x}{self.__clamp(g):02x}{self.__clamp(b):02x}".upper()

    def convert_to_excel(self, list_title: str = "Картина"):
        """
        Переводит изоображение в файл excel

        :return:
        """

        wb = Workbook()

        ws = wb.worksheets[0]
        ws.title = list_title

        for row_index, row in enumerate(self.__get_pixels(self.image_name), 1):
            for column_index, cell in enumerate(row, 1):
                color = Color(rgb=self.__get_hex_color(cell[0], cell[1], cell[2]))
                ws.cell(row=row_index, column=column_index).fill = PatternFill(bgColor=color,
                                                                               fgColor=color,
                                                                               fill_type="solid")

        for column_cells in ws.columns:
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = 3

        wb.save(filename=self.excel_name)
        print("Готово")


if __name__ == '__main__':
    item = ImageToExcel(image_name="img.png", excel_name="image.xlsx")
    item.convert_to_excel()
